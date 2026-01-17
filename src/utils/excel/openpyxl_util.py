import copy
import logging
import os
from pathlib import Path
import re
import openpyxl
from openpyxl.styles import Font, PatternFill

logger = logging.getLogger(__name__)


class OpenpyxlUtil:
    """
    Author: 김대광
    """

    _is_null_or_empty = "{} is null or empty"

    @classmethod
    def read_excel(
        cls, file_path: str, cell_names: list = None, is_decimal: bool = False
    ) -> list:
        """엑셀 파일 읽기

        Args:
            file_path (str): _description_
            cell_names (list, optional): _description_. Defaults to None.
            is_decimal (bool, optional): _description_. Defaults to False.

        Raises:
            ValueError: _description_
            FileNotFoundError: _description_

        Returns:
            _type_: _description_
        """
        if not file_path or not file_path.strip():
            raise ValueError(cls._is_null_or_empty.format("file_path"))

        path = Path(file_path)
        if not path.exists():
            raise FileNotFoundError(f"파일을 찾을 수 없습니다: {file_path}")

        cell_names = cell_names or []
        res_list = []

        try:
            wb = openpyxl.load_workbook(file_path, data_only=True)
            sheet = wb.active

            for row in sheet.iter_rows(min_row=2):
                row_map = {
                    name: cls.__get_cell_value(cell, is_decimal)
                    for name, cell in zip(cell_names, row)
                }

                res_list.append(row_map)

            return res_list
        except Exception as e:
            logger.error(f"Error: {e}")
            return []

    @classmethod
    def _get_cell_value(cls, cell, is_decimal: bool):
        val = cell.value

        if val is None:
            return ""

        if isinstance(val, (int, float)):
            return val if is_decimal else int(val)

        return val

    @staticmethod
    def write_excel(
        dest_path: str, file_name: str, contents_list: list, cell_titles: list
    ):
        """엑셀 파일 생성

        Args:
            dest_path (str): _description_
            file_name (str): _description_
            contents_list (list): _description_
            cell_titles (list): _description_

        Raises:
            ValueError: _description_

        Returns:
            _type_: _description_
        """
        if not all([dest_path, file_name, contents_list, cell_titles]):
            raise ValueError("모든 인자값은 필수이며 비어있을 수 없습니다.")

        wb = openpyxl.Workbook()
        ws = wb.active

        ws.append(cell_titles)

        header_font = Font(bold=True)
        header_fill = PatternFill(
            start_color="D3D3D3", end_color="D3D3D3", fill_type="solid"
        )

        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill

        for content in contents_list:
            row_data = [content.get(title, "") for title in cell_titles]
            ws.append(row_data)

        full_path = os.path.join(dest_path, file_name)

        try:
            wb.save(full_path)
            return True
        except Exception as e:
            logger.error(f"Error saving excel: {e}")
            return False

    @classmethod
    def write_excel_template(
        cls,
        template_file_path: str,
        dest_file_path: str,
        file_name: str,
        contents_dict: dict,
        list_key: str,
        list_start_row: int = 3,
    ):
        """템플릿 파일 이용해서 엑셀 파일 생성

        Args:
            template_file_path (str): _description_
            dest_file_path (str): _description_
            file_name (str): _description_
            contents_dict (dict): _description_
            list_key (str): _description_
            list_start_row (int, optional): _description_. Defaults to 3.

        Raises:
            ValueError: _description_
            ValueError: _description_
            ValueError: _description_
            ValueError: _description_

        Returns:
            _type_: _description_

        템플릿 작성:
            {{id}} | {{name}}
        """
        if not template_file_path or not template_file_path.strip():
            raise ValueError(cls._is_null_or_empty("template_file_path"))

        if not dest_file_path or not dest_file_path.strip():
            raise ValueError(cls._is_null_or_empty("dest_file_path"))

        if not file_name or not file_name.strip():
            raise ValueError(cls._is_null_or_empty("file_name"))

        if not contents_dict:
            raise ValueError(cls._is_null_or_empty("contents_dict"))

        if not list_key or not list_key.strip():
            raise ValueError(cls._is_null_or_empty("list_key"))

        try:
            wb = openpyxl.load_workbook(template_file_path)
            ws = wb.active

            # 일반 정보 치환
            cls._replace_global_placeholders(ws, contents_dict)

            # 리스트 정보 처리
            if list_key in contents_dict and isinstance(contents_dict[list_key], list):
                cls._render_list_data(ws, contents_dict[list_key], list_start_row)

            # 저장 로직
            os.makedirs(dest_file_path, exist_ok=True)
            save_path = os.path.join(dest_file_path, file_name)
            wb.save(save_path)
            return True
        except Exception as e:
            logger.error(f"Error occurred: {e}")
            return False

    @classmethod
    def _replace_global_placeholders(cls, ws, data):
        """시트 전체에서 리스트가 아닌 일반 단일 값들을 치환"""
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and "{{" in cell.value:
                    cell.value = cls._simple_replace(cell.value, data)

    @classmethod
    def _render_list_data(cls, ws, item_list, start_row):
        """리스트 데이터를 행 삽입 방식으로 렌더링"""
        if not item_list:
            return

        # 템플릿 행(서식 및 플레이스홀더) 캐싱
        template_cells = []

        for c in range(1, ws.max_column + 1):
            cell = ws.cell(row=start_row, column=c)
            template_cells.append(
                {
                    "value": cell.value,
                    "font": copy.copy(cell.font),
                    "border": copy.copy(cell.border),
                    "fill": copy.copy(cell.fill),
                    "alignment": copy.copy(cell.alignment),
                    "number_format": cell.number_format,
                }
            )

        # 데이터 개수만큼 행 삽입 (기존 데이터 밀어내기)
        # - 첫 번째 행은 이미 존재하므로 len-1 만큼 삽입
        if len(item_list) > 1:
            ws.insert_rows(start_row + 1, amount=len(item_list) - 1)

        # 데이터 기록
        for i, item_data in enumerate(item_list):
            current_row = start_row + i
            for c_idx, t_cell in enumerate(template_cells, 1):
                target_cell = ws.cell(row=current_row, column=c_idx)

                # 스타일 복사
                target_cell.font = t_cell["font"]
                target_cell.border = t_cell["border"]
                target_cell.fill = t_cell["fill"]
                target_cell.alignment = t_cell["alignment"]
                target_cell.number_format = t_cell["number_format"]

                # 값 치환
                val = t_cell["value"]
                if isinstance(val, str) and "{{" in val:
                    target_cell.value = cls._simple_replace(val, item_data)
                else:
                    target_cell.value = val

    @staticmethod
    def _simple_replace(text, data):
        matches = re.findall(r"\{\{(.*?)\}\}", text)
        for match in matches:
            key = match.strip()
            if key in data:
                text = text.replace(f"{{{{{match}}}}}", str(data[key]))
        return text
